import emoji
import pandas as pd
import openpyxl
from fpdf import FPDF
from datetime import datetime, date, timedelta


phone_model = ["SMALL Phone", "ReGuLaR Phone", "huge Phone"]
value = ["budget", "midrange", "expensive"]
cost = [429, 799, 999]
annual_interest_rate = 8


def main():
    # Ask user to choose a cellphone from the list OR what kind of budget he would like to spend on it (as input)
    while True:
        try:
            user_choice = input(f'Please choose your cell phone from criteria (exact item or by value) below:\n {", ".join(phone_model)}\n or\n {", ".join(value)}\nMy choise is: ').strip()
            if user_choice.lower().startswith("sm"):
                user_choice = phone_model[0]
            elif user_choice.lower().startswith("re"):
                user_choice = phone_model[1]
            elif user_choice.lower().startswith("hu"):
                user_choice = phone_model[2]
            elif user_choice.lower().startswith("b"):
                user_choice = value[0]
            elif user_choice.lower().startswith("m"):
                user_choice = value[1]
            elif user_choice.lower().startswith("e"):
                user_choice = value[2]
            else:
                raise ValueError
        except ValueError:
            print(emoji.emojize(":prohibited: -> Invalid input. Lets try again <- :prohibited:", language="alias"))
        else:
            break

    # Proceed with input to function_1, which tell user the price of phone by his choice (or program return exact item of the phone and price, if user choose budget range)
    # Ask again if user doesn't like his previous choise
    while True:
        try:
            user_selection = function_1(user_choice)
            confirmation = input(f"{user_selection}$\nDo you want to change your choice? (Yes/No): ").strip().lower()
            if confirmation.startswith("y"):
                while True:
                    try:
                        user_choice = input(f'You can choose another phone from criteria (exact item or by value):\n {", ".join(phone_model)}\n or\n {", ".join(value)}\nMy choise is: ').strip()
                        if user_choice.lower().startswith("sm"):
                            user_choice = phone_model[0]
                        elif user_choice.lower().startswith("re"):
                            user_choice = phone_model[1]
                        elif user_choice.lower().startswith("hu"):
                            user_choice = phone_model[2]
                        elif user_choice.lower().startswith("b"):
                            user_choice = value[0]
                        elif user_choice.lower().startswith("m"):
                            user_choice = value[1]
                        elif user_choice.lower().startswith("e"):
                            user_choice = value[2]
                        else:
                            raise ValueError
                        user_selection = function_1(user_choice)
                        print(f"{user_selection}$")
                    except ValueError:
                        print(emoji.emojize(":prohibited: -> Invalid input. Lets try again <- :prohibited:", language="alias"))
                    else:
                        break
            elif confirmation.startswith("n"):
                print("Good selection!")
                break
            else:
                raise ValueError
        except ValueError:
            print(emoji.emojize(":prohibited: -> Please choose Yes or No <- :prohibited:", language="alias"))
        else:
            break

    # After new choise ask does user wants to pay total amount of the price or put amount on monthly installments
    while True:
        try:
            payment = input('Do you want to pay the entire amount at once or arrange for monthly installments?\n Please choose "Total" or "Mothly Installments": ').strip().lower()
            # If user gonna select monthly installment - Proceed to function_2, which calculate loan for user's phone choice. And program will create xml file with calculation
            if payment.startswith("m") or payment.startswith("i"):
                while True:
                    try:
                        loan_term_months = int(input("Loan term (in months): "))
                        if loan_term_months < 1:
                            raise ValueError
                        break
                    except ValueError:
                        print(emoji.emojize(":prohibited: -> Type only digits. The loan is provided for more than one month <- :prohibited:", language="alias"))

                if user_choice == phone_model[0] or user_choice == value[0]:
                    loan_amount = cost[0]
                elif user_choice == phone_model[1] or user_choice == value[1]:
                    loan_amount = cost[1]
                else:
                    loan_amount = cost[2]

                schedule = function_2(loan_amount, loan_term_months, annual_interest_rate)

                df = pd.DataFrame(schedule)
                output_file = "installment_plan.xlsx"
                df.to_excel(output_file, index=False)
                print(pd.read_excel(output_file))
                print(emoji.emojize("Thanks for shopping :shopping_bags:  with us. We hope you will love your Phone :mobile_phone:. You will find the Installment Plan :chart_decreasing: in a new tab.", language="alias"))

            # If user gonna select payment with total amount - Proceed to function_n, which will return invoice message with expiration date. And program will create pdf invoice file
            elif payment.startswith("t"):
                if user_choice == phone_model[0] or user_choice == value[0]:
                    phone_to_invoice = phone_model[0]
                elif user_choice == phone_model[1] or user_choice == value[1]:
                    phone_to_invoice = phone_model[1]
                else:
                    phone_to_invoice = phone_model[2]

                voucer = function_n(phone_to_invoice)

                class PDF(FPDF):
                    def header(self):
                        self.set_font("helvetica", "B", 10)
                        pdf.set_text_color(128)
                        self.cell(130)
                        self.cell(60, 10, "B2B/B2C Phone Company", border=1, align="C")
                        self.ln(20)

                        self.set_font("helvetica", size=15)
                        pdf.set_text_color(0)
                        pdf.set_y(40)
                        self.cell(80)
                        self.cell(30, 10, "INVOICE", align="C")
                        self.ln(20)

                    def footer(self):
                        self.set_y(-15)
                        self.set_font("helvetica", "I", 8)
                        pdf.set_text_color(0, 100, 0)
                        self.cell(0, 10, f'Date of issue: {datetime.now().strftime("%Y-%m-%d %H:%M")}.')

                pdf = PDF(orientation="L", unit="mm", format="A5")
                pdf.add_page()
                pdf.set_font("Times", "IB", 12)
                pdf.set_text_color(58, 110, 122)
                pdf.set_y(75)
                pdf.cell(200, 10, voucer, align="C")
                pdf.output("invoice.pdf")
                print(emoji.emojize("Thanks for shopping :shopping_bags:  with us. We hope you will love your Phone :mobile_phone:. You will find the invoice :receipt: in a new tab.", language="alias"))

            else:
                raise ValueError
        except ValueError:
            print(emoji.emojize(":prohibited: -> Invalid input. Only Total or Installment could be chosen. <- :prohibited:", language="alias"))
        else:
            break

def function_1(choice):
    # Using user's input function should return cellphone from the list and it's cost (of just cost, if user chose model of his phone already)
    if choice == phone_model[0]:
        return f"Price for your phone is: {cost[0]}"
    elif choice == phone_model[1]:
        return f"Price for your phone is: {cost[1]}"
    elif choice == phone_model[2]:
        return f"Price for your phone is: {cost[2]}"
    elif choice == value[0]:
        return f"Price for your {phone_model[0]} is {cost[0]}"
    elif choice == value[1]:
        return f"Price for your {phone_model[1]} is {cost[1]}"
    else:
        return f"Price for your {phone_model[2]} is {cost[2]}"


def function_2(cost, loan_term_months, annual_interest_rate):
    # Function should return loan calculation
    monthly_interest_rate = annual_interest_rate / loan_term_months / 100

    schedule = []
    debt_balance = cost

    schedule.append({
        "Month": 1,
        "Monthly Payment (in $)": round(cost / loan_term_months, 2),
        "Balance of debt (in $)": round(debt_balance, 2),
        "Monthly Interest Rate": "{:.2%}".format(monthly_interest_rate),
        "Loan amount at interest (in $)": round(debt_balance * monthly_interest_rate, 2),
        "Loan (in $)": round((cost / loan_term_months) + (debt_balance * monthly_interest_rate), 2)
    })

    for month in range(2, loan_term_months + 1):
        monthly_payment = cost / loan_term_months
        debt_balance -= monthly_payment
        loan_amount_interest = debt_balance * monthly_interest_rate
        loan = monthly_payment + loan_amount_interest

        schedule.append({
            "Month": month,
            "Monthly Payment (in $)": round(monthly_payment, 2),
            "Balance of debt (in $)": round(debt_balance, 2),
            "Monthly Interest Rate": "{:.2%}".format(monthly_interest_rate),
            "Loan amount at interest (in $)": round(loan_amount_interest, 2),
            "Loan (in $)": round(loan, 2)
        })

    return schedule


def function_n(phone):
    # Function should return invoice message with expiration date
    expiration_date = date.today() + timedelta(days=14)
    invoice_message = f'This is a certificate for your new "{phone}". Validation only for 14 days till {expiration_date}.'
    return invoice_message


if __name__ == "__main__":
    main()
